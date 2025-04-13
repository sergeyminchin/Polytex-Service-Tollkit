import streamlit as st
import pandas as pd
from collections import defaultdict
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def run_app():
    st.title("🔁 Repeated Calls by Technician")

    uploaded_file = st.file_uploader("Upload Service Calls Excel File", type=["xlsx"])
    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file, engine='openpyxl')
        except Exception as e:
            st.error(f"Error reading Excel file: {e}")
            return

        call_id_column = None
        if "מס. קריאה" in df.columns:
            call_id_column = "מס. קריאה"
        elif "מספר קריאה" in df.columns:
            call_id_column = "מספר קריאה"
        else:
            st.error("The Excel file must contain either 'מס. קריאה' or 'מספר קריאה' columns.")
            return

        required_cols = [call_id_column, "ת. פתיחה", "מס' מכשיר", "לטיפול", "תאור תקלה", "תאור קוד פעולה"]
        if not all(col in df.columns for col in required_cols):
            st.error("❌ Missing one or more required columns.")
            st.write("🧾 Found:", df.columns.tolist())
            return

        df_relevant = df[required_cols].copy()
        df_relevant["ת. פתיחה"] = pd.to_datetime(df_relevant["ת. פתיחה"], errors="coerce")
        df_relevant = df_relevant.sort_values(by=["מס' מכשיר", "ת. פתיחה"])

        device_calls = defaultdict(list)
        for _, row in df_relevant.iterrows():
            device_id = row["מס' מכשיר"]
            call_id = row[call_id_column]
            open_date = row["ת. פתיחה"]
            technician = row["לטיפול"]
            fault_description = row["תאור תקלה"]
            action_description = row["תאור קוד פעולה"]

            if device_calls[device_id]:
                last_call = device_calls[device_id][-1]
                last_call_date = last_call["ת. פתיחה"]
                if (open_date - last_call_date).days <= 30:
                    last_call["קריאה חוזרת"].append({
                        "קריאה חוזרת": call_id,
                        "ת. פתיחה": open_date,
                        "לטיפול": technician,
                        "תאור תקלה": fault_description,
                        "תאור קוד פעולה": action_description
                    })

            device_calls[device_id].append({
                "קריאה ראשונה": call_id,
                "ת. פתיחה": open_date,
                "לטיפול": technician,
                "מס' מכשיר": device_id,
                "תאור תקלה": fault_description,
                "תאור קוד פעולה": action_description,
                "קריאה חוזרת": []
            })

        technician_data = defaultdict(list)
        total_calls = df_relevant.shape[0]
        total_repeats = 0

        for calls in device_calls.values():
            for call in calls:
                if call["קריאה חוזרת"]:
                    for repeat_call in call["קריאה חוזרת"]:
                        technician_data[call["לטיפול"]].append({
                            "קריאה ראשונה": call["קריאה ראשונה"],
                            "תאור תקלה (קריאה ראשונה)": call["תאור תקלה"],
                            "תאור קוד פעולה (קריאה ראשונה)": call["תאור קוד פעולה"],
                            "קריאה חוזרת": repeat_call["קריאה חוזרת"],
                            "תאור תקלה (קריאה חוזרת)": repeat_call["תאור תקלה"],
                            "תאור קוד פעולה (קריאה חוזרת)": repeat_call["תאור קוד פעולה"],
                            "מס' מכשיר": call["מס' מכשיר"]
                        })
                        total_repeats += 1

        technician_summary = {}
        for tech, records in technician_data.items():
            total_tech_calls = df_relevant[df_relevant["לטיפול"] == tech].shape[0]
            repeat_calls = len(records)
            repeat_call_percentage = (repeat_calls / total_tech_calls) * 100 if total_tech_calls else 0
            technician_summary[tech] = repeat_call_percentage

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for tech, records in technician_data.items():
                df_tech = pd.DataFrame(records)
                df_tech.to_excel(writer, sheet_name=tech[:31], index=False)

            summary_data = {
                "Total Calls": [total_calls],
                "Total Repeated Calls": [total_repeats],
                "Percentage of Repeated Calls": [f"{(total_repeats / total_calls) * 100:.2f}%" if total_calls else "0%"]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)

        output.seek(0)
        wb = load_workbook(output)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet_name != "Summary":
                percentage = technician_summary.get(sheet_name, 0)
                sheet.insert_rows(1)
                cell = sheet.cell(row=1, column=1)
                cell.value = f"Repeated Calls Percentage: {percentage:.2f}%"
                cell.font = Font(bold=True)

            for column_cells in sheet.columns:
                max_length = 0
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.success("📊 Analysis complete. Download the Excel file below.")
        st.download_button(
            label="📥 Download Excel File with Technician Tabs",
            data=final_output,
            file_name="repeated_calls_by_technician_tabs.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
